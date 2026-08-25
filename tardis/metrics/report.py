"""Strict, atomic single- and three-dataset metric reports."""

from __future__ import annotations

import csv
import os
import tempfile
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from math import fsum, isfinite
from pathlib import Path
from typing import Literal, Self, cast

from openpyxl import Workbook  # type: ignore[import-untyped]

TestDataset = Literal["dataverse_test", "openvid_test", "seedance_test"]
ReportDataset = Literal[
    "dataverse_test",
    "openvid_test",
    "seedance_test",
    "average",
]

TEST_DATASETS: tuple[TestDataset, ...] = (
    "dataverse_test",
    "openvid_test",
    "seedance_test",
)
METRIC_FIELDS: tuple[str, ...] = ("tc", "lpips", "fvd", "fid", "clipscore", "ssim")
REPORT_FIELDS: tuple[str, ...] = ("dataset", *METRIC_FIELDS)


@dataclass(frozen=True, slots=True)
class MetricValues:
    """The six finite values reported for one dataset."""

    tc: float
    lpips: float
    fvd: float
    fid: float
    clipscore: float
    ssim: float

    def __post_init__(self) -> None:
        for metric, raw_value in zip(METRIC_FIELDS, self.as_tuple(), strict=True):
            if isinstance(raw_value, bool):
                raise ValueError(f"{metric} must be a finite number")
            try:
                value = float(raw_value)
            except (TypeError, ValueError) as error:
                raise ValueError(f"{metric} must be a finite number") from error
            if not isfinite(value):
                raise ValueError(f"{metric} must be finite")
            object.__setattr__(self, metric, value)

    @classmethod
    def from_mapping(cls, dataset: str, values: Mapping[str, float]) -> Self:
        if set(values) != set(METRIC_FIELDS):
            fields = ", ".join(METRIC_FIELDS)
            raise ValueError(f"{dataset} metrics must contain exactly: {fields}")
        converted: dict[str, float] = {}
        for metric in METRIC_FIELDS:
            try:
                value = float(values[metric])
            except (TypeError, ValueError) as error:
                raise ValueError(f"{dataset}.{metric} must be a finite number") from error
            if not isfinite(value):
                raise ValueError(f"{dataset}.{metric} must be finite")
            converted[metric] = value
        return cls(**converted)

    def as_tuple(self) -> tuple[float, ...]:
        return (self.tc, self.lpips, self.fvd, self.fid, self.clipscore, self.ssim)


@dataclass(frozen=True, slots=True)
class MetricReportRow:
    """One stable report row."""

    dataset: ReportDataset
    metrics: MetricValues

    def __post_init__(self) -> None:
        if self.dataset not in (*TEST_DATASETS, "average"):
            raise ValueError(f"invalid report dataset: {self.dataset!r}")
        if not isinstance(self.metrics, MetricValues):
            raise TypeError("metrics must be MetricValues")

    def as_tuple(self) -> tuple[str | float, ...]:
        return (self.dataset, *self.metrics.as_tuple())


@dataclass(frozen=True, slots=True)
class ThreeDatasetMetricReport:
    """Validated test rows followed by their equal-weight arithmetic mean."""

    rows: tuple[MetricReportRow, MetricReportRow, MetricReportRow, MetricReportRow]

    def __post_init__(self) -> None:
        if not isinstance(self.rows, tuple) or len(self.rows) != 4:
            raise ValueError("report must contain exactly four rows")
        if any(not isinstance(row, MetricReportRow) for row in self.rows):
            raise TypeError("report rows must be MetricReportRow instances")
        expected_order = (*TEST_DATASETS, "average")
        actual_order = tuple(row.dataset for row in self.rows)
        if actual_order != expected_order:
            raise ValueError(f"report row order must be {expected_order}")
        expected_average = tuple(
            _finite_mean(tuple(row.metrics.as_tuple()[index] for row in self.rows[:3]))
            for index in range(len(METRIC_FIELDS))
        )
        if self.rows[-1].metrics.as_tuple() != expected_average:
            raise ValueError("average row must equal the three-dataset arithmetic mean")

    @classmethod
    def from_mapping(cls, results: Mapping[str, Mapping[str, float]]) -> Self:
        if set(results) != set(TEST_DATASETS):
            datasets = ", ".join(TEST_DATASETS)
            raise ValueError(f"report must contain exactly {datasets}")

        test_rows = (
            MetricReportRow(
                "dataverse_test",
                MetricValues.from_mapping("dataverse_test", results["dataverse_test"]),
            ),
            MetricReportRow(
                "openvid_test",
                MetricValues.from_mapping("openvid_test", results["openvid_test"]),
            ),
            MetricReportRow(
                "seedance_test",
                MetricValues.from_mapping("seedance_test", results["seedance_test"]),
            ),
        )
        average_values = tuple(
            _finite_mean(tuple(row.metrics.as_tuple()[index] for row in test_rows))
            for index in range(len(METRIC_FIELDS))
        )
        average = MetricValues(*average_values)
        first, second, third = test_rows
        return cls((first, second, third, MetricReportRow("average", average)))

    def write_csv(self, destination: Path) -> None:
        """Atomically write the four report rows as UTF-8 CSV."""

        def write(temporary: Path) -> None:
            with temporary.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(REPORT_FIELDS)
                writer.writerows(row.as_tuple() for row in self.rows)
                handle.flush()
                os.fsync(handle.fileno())

        _atomic_write(Path(destination), write)

    def write_xlsx(self, destination: Path) -> None:
        """Atomically write the same four rows to one XLSX worksheet."""

        def write(temporary: Path) -> None:
            workbook = Workbook(write_only=True)
            worksheet = workbook.create_sheet(title="metrics")
            worksheet.append(REPORT_FIELDS)
            for row in self.rows:
                worksheet.append(row.as_tuple())
            workbook.save(temporary)
            with temporary.open("rb") as handle:
                os.fsync(handle.fileno())

        _atomic_write(Path(destination), write)


@dataclass(frozen=True, slots=True)
class SingleDatasetMetricReport:
    """One selected test dataset's six metrics, without cross-dataset averaging."""

    row: MetricReportRow

    def __post_init__(self) -> None:
        if self.row.dataset not in TEST_DATASETS:
            raise ValueError("single-dataset report requires one canonical test dataset")

    @classmethod
    def from_mapping(cls, dataset: str, values: Mapping[str, float]) -> Self:
        if dataset not in TEST_DATASETS:
            raise ValueError(f"invalid test dataset: {dataset!r}")
        typed_dataset = cast(TestDataset, dataset)
        return cls(MetricReportRow(typed_dataset, MetricValues.from_mapping(dataset, values)))

    def write_csv(self, destination: Path) -> None:
        def write(temporary: Path) -> None:
            with temporary.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(REPORT_FIELDS)
                writer.writerow(self.row.as_tuple())
                handle.flush()
                os.fsync(handle.fileno())

        _atomic_write(Path(destination), write)

    def write_xlsx(self, destination: Path) -> None:
        def write(temporary: Path) -> None:
            workbook = Workbook(write_only=True)
            worksheet = workbook.create_sheet(title="metrics")
            worksheet.append(REPORT_FIELDS)
            worksheet.append(self.row.as_tuple())
            workbook.save(temporary)
            with temporary.open("rb") as handle:
                os.fsync(handle.fileno())

        _atomic_write(Path(destination), write)


def _atomic_write(destination: Path, write: Callable[[Path], None]) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{destination.name}.",
            suffix=".tmp",
            dir=destination.parent,
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
        write(temporary)
        os.replace(temporary, destination)
    except BaseException:
        if temporary is not None:
            temporary.unlink(missing_ok=True)
        raise


def _finite_mean(values: tuple[float, ...]) -> float:
    mean = fsum(value / len(values) for value in values)
    if not isfinite(mean):
        raise ValueError("metric average must be finite")
    return mean
