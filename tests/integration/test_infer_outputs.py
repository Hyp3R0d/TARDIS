from __future__ import annotations

import argparse
import csv
import json
import weakref
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest
import torch
from openpyxl import load_workbook

from tardis.data.contracts import VideoRecord
from tardis.data.dataset import ClipBatch
from tardis.utils.video_io import probe_video


class FakeContext:
    rank = 0
    local_rank = 0
    world_size = 1
    device = torch.device("cpu")
    is_main = True

    def initialize(self) -> None:
        pass

    def barrier(self) -> None:
        pass

    def close(self) -> None:
        pass


class FakeMetricSuite:
    provenance_ids = {
        name: f"fake-{name}"
        for name in (
            "tc",
            "lpips",
            "fvd",
            "fid",
            "clipscore",
            "ssim",
        )
    }

    def __init__(self) -> None:
        self.count = 0
        self.reduced = False

    def update(self, generated: torch.Tensor, reference: torch.Tensor, prompt: str) -> None:
        assert generated.ndim == reference.ndim == 4
        assert generated.shape == reference.shape
        assert prompt
        self.count += 1

    def compute(self) -> dict[str, dict[str, float]]:
        values = {
            "tc": float(self.count),
            "lpips": float(self.count + 1),
            "fvd": float(self.count + 2),
            "fid": float(self.count + 3),
            "clipscore": float(self.count + 4),
            "ssim": float(self.count + 5),
        }
        return {"macro": values, "micro": dict(values)}

    def state_dict(self) -> dict[str, object]:
        return {"count": self.count}

    def load_state_dict(self, state: dict[str, object]) -> None:
        self.count = int(state["count"])

    def all_reduce(self) -> None:
        self.reduced = True

    def reset(self) -> None:
        self.count = 0
        self.reduced = False


class FakeModel:
    def __init__(self, *, interrupt_after: int | None = None) -> None:
        self.interrupt_after = interrupt_after
        self.calls: list[str] = []

    def eval(self) -> FakeModel:
        return self

    def generate(
        self,
        prompts: list[str],
        num_frames: int,
        fps: int,
        generator: torch.Generator,
    ) -> SimpleNamespace:
        del fps, generator
        prompt = prompts[0]
        self.calls.append(prompt)
        if self.interrupt_after is not None and len(self.calls) > self.interrupt_after:
            raise KeyboardInterrupt
        video = torch.zeros(1, num_frames, 3, 16, 16)
        video[:, :, 0] = min(len(self.calls) / 10, 1.0)
        return SimpleNamespace(video=video)


class LightweightLPIPS:
    provenance_id = "integration/lpips"

    def __call__(self, generated: torch.Tensor, reference: torch.Tensor) -> torch.Tensor:
        return (generated - reference).abs().mean(dim=(1, 2, 3))


class LightweightFrameFeatures:
    provenance_id = "integration/fid"
    feature_dim = 2

    def __call__(self, video: torch.Tensor) -> torch.Tensor:
        means = video.mean(dim=(1, 2, 3))
        return torch.stack((means, means.square()), dim=1)


class LightweightVideoFeatures:
    provenance_id = "integration/fvd"
    feature_dim = 2

    def __call__(self, video: torch.Tensor) -> torch.Tensor:
        mean = video.mean()
        return torch.stack((mean, mean.square())).reshape(1, 2)


class LightweightCLIPFeatures:
    provenance_id = "integration/clip"
    feature_dim = 2

    def encode_video(self, video: torch.Tensor) -> torch.Tensor:
        means = video.mean(dim=(1, 2, 3))
        return torch.stack((torch.ones_like(means), means), dim=1)

    def encode_text(self, prompt: str) -> torch.Tensor:
        assert prompt
        return torch.tensor([[1.0, 0.0]])


def _real_metric_suite() -> Any:
    from tardis.metrics.frechet import FIDMetric, FVDMetric
    from tardis.metrics.paired import CLIPScoreMetric, LPIPSMetric
    from tardis.metrics.suite import MetricSuite

    return MetricSuite(
        lpips=LPIPSMetric(LightweightLPIPS()),
        fid=FIDMetric(LightweightFrameFeatures()),
        fvd=FVDMetric(LightweightVideoFeatures()),
        clipscore=CLIPScoreMetric(LightweightCLIPFeatures()),
    )


@dataclass
class FakeLoaders:
    test: dict[str, list[ClipBatch]]


def _batch(source: str, index: int) -> ClipBatch:
    record_id = f"{source}-{index}"
    return ClipBatch(
        prompts=[f"prompt {record_id}"],
        video=torch.zeros(1, 2, 3, 16, 16),
        sources=(source,),
        record_ids=(record_id,),
        sample_seeds=(100 + index,),
    )


def _loaders(source: str = "dataverse", count: int = 6) -> FakeLoaders:
    return FakeLoaders(test={source: [_batch(source, index) for index in range(count)]})


def _runtime(model: FakeModel, checkpoint: Path) -> SimpleNamespace:
    return SimpleNamespace(
        model=model,
        metric_suite=FakeMetricSuite(),
        checkpoint=SimpleNamespace(path=checkpoint, sha256="a" * 64, used_ema=True),
        dataset_sources=("dataverse",),
        device=torch.device("cpu"),
    )


def _services(
    model: FakeModel,
    checkpoint: Path,
    *,
    loaders: FakeLoaders | None = None,
) -> Any:
    from tardis.cli.infer import InferServices

    def runtime_builder(args: object, **kwargs: object) -> SimpleNamespace:
        del args
        assert kwargs["use_ema"] is True
        return _runtime(model, checkpoint)

    return InferServices(
        context_factory=lambda _device_type: FakeContext(),
        runtime_builder=runtime_builder,
        loader_builder=lambda args, _context: (
            _loaders(str(args.dataset)) if loaders is None else loaders
        ),
    )


def test_infer_writes_complete_resumable_single_dataset_artifacts(tmp_path: Path) -> None:
    from tardis.cli.infer import parse_args, run_inference

    checkpoint = tmp_path / "best.pt"
    checkpoint.write_bytes(b"fake checkpoint")
    output_root = tmp_path / "outputs"
    first_model = FakeModel(interrupt_after=3)
    args = parse_args(
        [
            "--checkpoint",
            str(checkpoint),
            "--output-root",
            str(output_root),
            "--device",
            "cpu",
            "--precision",
            "fp32",
            "--test-size",
            "6",
            "--num-frames",
            "2",
            "--height",
            "16",
            "--width",
            "16",
            "--num-workers",
            "0",
        ]
    )

    with pytest.raises(KeyboardInterrupt):
        run_inference(args, services=_services(first_model, checkpoint))

    run_dir = next((output_root / "infer" / "dataverse").iterdir())
    state = torch.load(
        run_dir / "rank_0000" / "dataverse_test.metrics.pt",
        map_location="cpu",
        weights_only=True,
    )
    assert state["completed_ids"] == ["dataverse-0", "dataverse-1", "dataverse-2"]

    resumed_model = FakeModel()
    resumed_args = parse_args([*map(str, args_to_argv(args)), "--resume-output", str(run_dir)])
    result = run_inference(resumed_args, services=_services(resumed_model, checkpoint))

    assert result == run_dir
    assert resumed_model.calls[:3] == [
        "prompt dataverse-3",
        "prompt dataverse-4",
        "prompt dataverse-5",
    ]
    completed = [
        json.loads(line)
        for line in (run_dir / "rank_0000" / "completed.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
    ]
    assert {item["record_id"] for item in completed} == {f"dataverse-{index}" for index in range(6)}

    with (run_dir / "metrics.csv").open(newline="", encoding="utf-8") as handle:
        csv_rows = list(csv.DictReader(handle))
    assert [row["dataset"] for row in csv_rows] == ["dataverse_test"]
    workbook = load_workbook(run_dir / "metrics.xlsx", read_only=True)
    xlsx_rows = list(workbook["metrics"].iter_rows(values_only=True))
    workbook.close()
    assert [row[0] for row in xlsx_rows[1:]] == ["dataverse_test"]
    assert len(xlsx_rows) == 2

    with (run_dir / "per_video_details.csv").open(newline="", encoding="utf-8") as handle:
        details = list(csv.DictReader(handle))
    assert {row["record_id"] for row in details} == {f"dataverse-{index}" for index in range(6)}
    assert len((run_dir / "per_video_details.jsonl").read_text().splitlines()) == 6
    assert (run_dir / "failures.jsonl").is_file()
    assert (run_dir / "failures.jsonl").read_text(encoding="utf-8") == ""
    for report in ("manifest.json", "resources.json", "latency.json"):
        assert (run_dir / report).is_file()

    showcases = sorted((run_dir / "showcases").glob("*.mp4"))
    assert len(showcases) == 5
    assert {path.name.split("__", 1)[0] for path in showcases} == {"dataverse"}
    assert all(probe_video(path).playable for path in showcases)
    assert not list(run_dir.rglob("*.png"))
    assert not list(run_dir.rglob("*.jpg"))
    assert not list(run_dir.rglob("*.jpeg"))


@pytest.mark.parametrize(
    ("field", "changed_value"),
    [
        ("use_ema", False),
        ("precision", "bf16"),
        ("compile_model", True),
        ("deterministic", True),
        ("pretrained_model", "different/prior"),
        ("active_ratio", 0.5),
        ("validation_size", 1),
    ],
)
def test_infer_resume_rejects_generation_or_metric_mode_change(
    tmp_path: Path,
    field: str,
    changed_value: object,
) -> None:
    import tardis.cli.infer as infer_module

    args = infer_module.parse_args(["--device", "cpu", "--precision", "fp32"])
    state_path = tmp_path / "dataverse_test.metrics.pt"
    infer_module._save_progress(
        state_path,
        suite=FakeMetricSuite(),
        source="dataverse",
        args=args,
        context=FakeContext(),
        checkpoint_sha="d" * 64,
        records=[],
    )
    changed = argparse.Namespace(**vars(args))
    setattr(changed, field, changed_value)

    with pytest.raises(ValueError, match="incompatible infer resume state field: settings"):
        infer_module._load_progress(
            state_path,
            suite=FakeMetricSuite(),
            source="dataverse",
            args=changed,
            context=FakeContext(),
            checkpoint_sha="d" * 64,
        )


def test_infer_reconcile_truncates_torn_journal_tail_from_atomic_state(
    tmp_path: Path,
) -> None:
    import tardis.cli.infer as infer_module

    records = [
        {"source": "dataverse", "record_id": record_id}
        for record_id in ("dataverse-0", "dataverse-1", "dataverse-2")
    ]
    journal = tmp_path / "completed.jsonl"
    journal.write_bytes(
        (json.dumps(records[0]) + "\n" + json.dumps(records[1]) + "\n").encode()
        + b'{"record_id":"dataverse-2"'
    )

    infer_module._reconcile_journal(journal, records)

    reconciled = [json.loads(line) for line in journal.read_text(encoding="utf-8").splitlines()]
    assert [record["record_id"] for record in reconciled] == [
        "dataverse-0",
        "dataverse-1",
        "dataverse-2",
    ]
    assert journal.read_bytes().endswith(b"\n")


def test_infer_reconcile_rejects_corrupt_journal_middle(tmp_path: Path) -> None:
    import tardis.cli.infer as infer_module

    records = [
        {"source": "dataverse", "record_id": record_id}
        for record_id in ("dataverse-0", "dataverse-1")
    ]
    journal = tmp_path / "completed.jsonl"
    journal.write_text(
        json.dumps(records[0]) + "\n" + '{"broken":' + "\n" + json.dumps(records[1]) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(json.JSONDecodeError):
        infer_module._reconcile_journal(journal, records)


def test_infer_records_decode_failure_and_continues_source(tmp_path: Path) -> None:
    from tardis.cli.infer import parse_args, run_inference
    from tardis.data.dataset import BenchmarkFailure

    checkpoint = tmp_path / "best.pt"
    checkpoint.write_bytes(b"fake checkpoint")
    model = FakeModel()
    loaders = _loaders()
    failed_record = VideoRecord(
        id="dataverse-1",
        caption="prompt dataverse-1",
        media_locator="https://hf-mirror.com/failure.mp4",
        source="dataverse",
    )
    loaders.test["dataverse"][1] = BenchmarkFailure(
        record=failed_record,
        sample_seed=101,
        error_type="VideoDecodeError",
        error_message="invalid media payload",
    )
    args = parse_args(
        [
            "--checkpoint",
            str(checkpoint),
            "--output-root",
            str(tmp_path / "outputs"),
            "--device",
            "cpu",
            "--precision",
            "fp32",
            "--test-size",
            "6",
            "--num-frames",
            "2",
            "--height",
            "16",
            "--width",
            "16",
            "--num-workers",
            "0",
        ]
    )

    run_dir = run_inference(
        args,
        services=_services(model, checkpoint, loaders=loaders),
    )

    failures = [
        json.loads(line)
        for line in (run_dir / "failures.jsonl").read_text(encoding="utf-8").splitlines()
    ]
    assert [(item["record_id"], item["error_type"]) for item in failures] == [
        ("dataverse-1", "VideoDecodeError")
    ]
    assert "prompt dataverse-5" in model.calls
    assert len(model.calls) == 10
    assert len(list((run_dir / "showcases").glob("*.mp4"))) == 5


def test_infer_resume_retries_failed_record_and_replaces_failure_journal(
    tmp_path: Path,
) -> None:
    from tardis.cli.infer import parse_args, run_inference
    from tardis.data.dataset import BenchmarkFailure

    checkpoint = tmp_path / "best.pt"
    checkpoint.write_bytes(b"fake checkpoint")
    output_root = tmp_path / "outputs"
    failed_loaders = _loaders()
    failed_loaders.test["dataverse"][1] = BenchmarkFailure(
        record=VideoRecord(
            id="dataverse-1",
            caption="prompt dataverse-1",
            media_locator="https://hf-mirror.com/failure.mp4",
            source="dataverse",
        ),
        sample_seed=101,
        error_type="VideoDecodeError",
        error_message="transient decode failure",
    )
    args = parse_args(
        [
            "--checkpoint",
            str(checkpoint),
            "--output-root",
            str(output_root),
            "--device",
            "cpu",
            "--precision",
            "fp32",
            "--test-size",
            "6",
            "--num-frames",
            "2",
            "--height",
            "16",
            "--width",
            "16",
            "--num-workers",
            "0",
        ]
    )
    run_dir = run_inference(
        args,
        services=_services(FakeModel(), checkpoint, loaders=failed_loaders),
    )

    resumed_model = FakeModel()
    resumed_args = parse_args([*map(str, args_to_argv(args)), "--resume-output", str(run_dir)])
    run_inference(
        resumed_args,
        services=_services(resumed_model, checkpoint, loaders=_loaders()),
    )

    state = torch.load(
        run_dir / "rank_0000" / "dataverse_test.metrics.pt",
        map_location="cpu",
        weights_only=True,
    )
    assert state["metric_suite"]["count"] == 6
    assert len(state["records"]) == 6
    assert all(record["status"] == "completed" for record in state["records"])
    journal = [
        json.loads(line)
        for line in (run_dir / "rank_0000" / "completed.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
    ]
    assert len(journal) == 6
    assert len({record["record_id"] for record in journal}) == 6
    assert all(record["status"] == "completed" for record in journal)
    assert resumed_model.calls[0] == "prompt dataverse-1"
    assert (run_dir / "failures.jsonl").read_text(encoding="utf-8") == ""


def test_infer_passes_unbatched_single_video_tensors_to_real_metric_suite(
    tmp_path: Path,
) -> None:
    import tardis.cli.infer as infer_module

    args = infer_module.parse_args(
        [
            "--device",
            "cpu",
            "--precision",
            "fp32",
            "--test-size",
            "1",
            "--num-frames",
            "2",
            "--height",
            "16",
            "--width",
            "16",
            "--num-workers",
            "0",
        ]
    )
    context = FakeContext()
    context.is_main = False
    rank_dir = tmp_path / "rank_0000"
    output_dir = tmp_path / "output"
    rank_dir.mkdir()
    output_dir.mkdir()
    records: list[dict[str, object]] = []

    infer_module._evaluate_source(
        loader=[_batch("dataverse", 0)],
        model=FakeModel(),
        suite=_real_metric_suite(),
        source="dataverse",
        dataset="dataverse_test",
        args=args,
        context=context,
        runtime=SimpleNamespace(
            device=torch.device("cpu"),
            checkpoint=SimpleNamespace(sha256="c" * 64),
        ),
        rank_dir=rank_dir,
        output_dir=output_dir,
        records=records,
        make_generator=lambda seed, device: torch.Generator(device=device).manual_seed(seed),
    )

    assert [record["status"] for record in records] == ["completed"]


def test_infer_oom_leaves_sample_uncompleted_and_retryable(tmp_path: Path) -> None:
    import tardis.cli.infer as infer_module

    class OOMModel:
        def generate(self, *args: object, **kwargs: object) -> None:
            del args, kwargs
            raise torch.cuda.OutOfMemoryError("test oom")

    args = infer_module.parse_args(
        [
            "--device",
            "cpu",
            "--precision",
            "fp32",
            "--test-size",
            "1",
            "--num-frames",
            "2",
            "--height",
            "16",
            "--width",
            "16",
            "--num-workers",
            "0",
        ]
    )
    context = FakeContext()
    context.is_main = False
    rank_dir = tmp_path / "rank_0000"
    output_dir = tmp_path / "output"
    rank_dir.mkdir()
    output_dir.mkdir()
    state_path = rank_dir / "dataverse_test.metrics.pt"
    suite = FakeMetricSuite()
    records: list[dict[str, object]] = []
    infer_module._save_progress(
        state_path,
        suite=suite,
        source="dataverse",
        args=args,
        context=context,
        checkpoint_sha="c" * 64,
        records=records,
    )

    with pytest.raises(torch.cuda.OutOfMemoryError, match="test oom"):
        infer_module._evaluate_source(
            loader=[_batch("dataverse", 0)],
            model=OOMModel(),
            suite=suite,
            source="dataverse",
            dataset="dataverse_test",
            args=args,
            context=context,
            runtime=SimpleNamespace(
                device=torch.device("cpu"),
                checkpoint=SimpleNamespace(sha256="c" * 64),
            ),
            rank_dir=rank_dir,
            output_dir=output_dir,
            records=records,
            make_generator=lambda seed, device: torch.Generator(device=device).manual_seed(seed),
        )

    state = torch.load(state_path, map_location="cpu", weights_only=True)
    assert records == []
    assert state["completed_ids"] == []
    assert not (rank_dir / "completed.jsonl").exists()


def test_infer_releases_complete_output_before_next_generate(tmp_path: Path) -> None:
    import tardis.cli.infer as infer_module

    @dataclass
    class GeneratedOutput:
        video: torch.Tensor

    class WeakrefModel:
        def __init__(self) -> None:
            self.previous: weakref.ReferenceType[GeneratedOutput] | None = None
            self.previous_reachable_at_generate: list[bool] = []

        def generate(
            self,
            prompts: list[str],
            num_frames: int,
            fps: int,
            generator: torch.Generator,
        ) -> GeneratedOutput:
            del prompts, fps, generator
            if self.previous is not None:
                self.previous_reachable_at_generate.append(self.previous() is not None)
            output = GeneratedOutput(torch.zeros(1, num_frames, 3, 16, 16))
            self.previous = weakref.ref(output)
            return output

    args = infer_module.parse_args(
        [
            "--device",
            "cpu",
            "--precision",
            "fp32",
            "--test-size",
            "2",
            "--num-frames",
            "2",
            "--height",
            "16",
            "--width",
            "16",
            "--num-workers",
            "0",
        ]
    )
    context = FakeContext()
    context.is_main = False
    rank_dir = tmp_path / "rank_0000"
    output_dir = tmp_path / "output"
    rank_dir.mkdir()
    output_dir.mkdir()
    model = WeakrefModel()

    infer_module._evaluate_source(
        loader=[_batch("dataverse", 0), _batch("dataverse", 1)],
        model=model,
        suite=FakeMetricSuite(),
        source="dataverse",
        dataset="dataverse_test",
        args=args,
        context=context,
        runtime=SimpleNamespace(
            device=torch.device("cpu"),
            checkpoint=SimpleNamespace(sha256="c" * 64),
        ),
        rank_dir=rank_dir,
        output_dir=output_dir,
        records=[],
        make_generator=lambda seed, device: torch.Generator(device=device).manual_seed(seed),
    )

    assert model.previous_reachable_at_generate == [False]


def args_to_argv(args: object) -> list[object]:
    return [
        "--dataset",
        args.dataset,
        "--checkpoint",
        args.checkpoint,
        "--output-root",
        args.output_root,
        "--device",
        args.device,
        "--precision",
        args.precision,
        "--test-size",
        args.test_size,
        "--num-frames",
        args.num_frames,
        "--height",
        args.height,
        "--width",
        args.width,
        "--num-workers",
        args.num_workers,
    ]
