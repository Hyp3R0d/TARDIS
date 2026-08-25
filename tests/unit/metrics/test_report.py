from __future__ import annotations

import csv
import math
import sys
from pathlib import Path
from typing import Any, cast

import openpyxl
import pytest

from tardis.metrics.report import (
    METRIC_FIELDS,
    REPORT_FIELDS,
    TEST_DATASETS,
    MetricReportRow,
    MetricValues,
    ThreeDatasetMetricReport,
)


def _results() -> dict[str, dict[str, float]]:
    return {
        "dataverse_test": {
            "tc": 0.10,
            "lpips": 0.20,
            "fvd": 10.0,
            "fid": 20.0,
            "clipscore": 0.30,
            "ssim": 0.40,
        },
        "openvid_test": {
            "tc": 0.20,
            "lpips": 0.30,
            "fvd": 20.0,
            "fid": 30.0,
            "clipscore": 0.40,
            "ssim": 0.50,
        },
        "seedance_test": {
            "tc": 0.30,
            "lpips": 0.40,
            "fvd": 30.0,
            "fid": 40.0,
            "clipscore": 0.50,
            "ssim": 0.60,
        },
    }


@pytest.mark.unit
def test_report_requires_three_test_rows_and_builds_equal_weight_average() -> None:
    report = ThreeDatasetMetricReport.from_mapping(_results())

    assert TEST_DATASETS == ("dataverse_test", "openvid_test", "seedance_test")
    assert METRIC_FIELDS == ("tc", "lpips", "fvd", "fid", "clipscore", "ssim")
    assert ("dataset", *METRIC_FIELDS) == REPORT_FIELDS
    assert [row.dataset for row in report.rows] == [*TEST_DATASETS, "average"]
    assert report.rows[-1].metrics.tc == pytest.approx(0.20)
    assert report.rows[-1].metrics.lpips == pytest.approx(0.30)
    assert report.rows[-1].metrics.fvd == pytest.approx(20.0)
    assert report.rows[-1].metrics.fid == pytest.approx(30.0)
    assert report.rows[-1].metrics.clipscore == pytest.approx(0.40)
    assert report.rows[-1].metrics.ssim == pytest.approx(0.50)


@pytest.mark.unit
@pytest.mark.parametrize(
    "bad_dataset",
    ["dataverse_validation", "openvid_validation", "seedance_validation", "average"],
)
def test_report_rejects_validation_or_precomputed_average_rows(bad_dataset: str) -> None:
    values = _results()
    values[bad_dataset] = values.pop("dataverse_test")

    with pytest.raises(ValueError, match="exactly.*dataverse_test.*openvid_test.*seedance_test"):
        ThreeDatasetMetricReport.from_mapping(values)


@pytest.mark.unit
@pytest.mark.parametrize("metric", METRIC_FIELDS)
@pytest.mark.parametrize("bad_value", [math.nan, math.inf, -math.inf])
def test_report_rejects_non_finite_metric_values(metric: str, bad_value: float) -> None:
    values = _results()
    values["openvid_test"][metric] = bad_value

    with pytest.raises(ValueError, match=rf"openvid_test\.{metric}.*finite"):
        ThreeDatasetMetricReport.from_mapping(values)


@pytest.mark.unit
def test_report_rejects_missing_or_extra_metrics() -> None:
    missing = _results()
    del missing["dataverse_test"]["ssim"]
    with pytest.raises(ValueError, match="exactly.*tc.*ssim"):
        ThreeDatasetMetricReport.from_mapping(missing)

    extra = _results()
    extra["seedance_test"]["psnr"] = 42.0
    with pytest.raises(ValueError, match="exactly.*tc.*ssim"):
        ThreeDatasetMetricReport.from_mapping(extra)


@pytest.mark.unit
def test_csv_and_xlsx_have_identical_stable_rows(tmp_path: Path) -> None:
    report = ThreeDatasetMetricReport.from_mapping(_results())
    csv_path = tmp_path / "metrics.csv"
    xlsx_path = tmp_path / "metrics.xlsx"

    report.write_csv(csv_path)
    report.write_xlsx(xlsx_path)

    with csv_path.open(newline="", encoding="utf-8") as handle:
        csv_rows = list(csv.reader(handle))
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        xlsx_rows = [list(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()

    assert csv_rows[0] == list(REPORT_FIELDS)
    assert xlsx_rows[0] == list(REPORT_FIELDS)
    assert [row[0] for row in csv_rows[1:]] == [*TEST_DATASETS, "average"]
    assert [row[0] for row in xlsx_rows[1:]] == [*TEST_DATASETS, "average"]
    assert len(csv_rows) == len(xlsx_rows) == 5
    for csv_row, xlsx_row in zip(csv_rows[1:], xlsx_rows[1:], strict=True):
        assert csv_row[0] == xlsx_row[0]
        assert [float(value) for value in csv_row[1:]] == pytest.approx(xlsx_row[1:])


@pytest.mark.unit
def test_large_finite_metrics_have_finite_consistent_csv_and_xlsx_average(
    tmp_path: Path,
) -> None:
    near_limit = sys.float_info.max / 2.0
    values = {
        dataset: {metric: near_limit for metric in METRIC_FIELDS} for dataset in TEST_DATASETS
    }

    report = ThreeDatasetMetricReport.from_mapping(values)

    average = report.rows[-1].metrics.as_tuple()
    assert all(math.isfinite(value) for value in average)
    assert average == pytest.approx((near_limit,) * len(METRIC_FIELDS), rel=1e-15)

    csv_path = tmp_path / "large.csv"
    xlsx_path = tmp_path / "large.xlsx"
    report.write_csv(csv_path)
    report.write_xlsx(xlsx_path)
    with csv_path.open(newline="", encoding="utf-8") as handle:
        csv_average = tuple(float(value) for value in list(csv.reader(handle))[-1][1:])
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        xlsx_average = tuple(float(value) for value in list(workbook.active.values)[-1][1:])
    finally:
        workbook.close()

    assert all(math.isfinite(value) for value in (*csv_average, *xlsx_average))
    assert csv_average == xlsx_average == average


@pytest.mark.unit
def test_metric_values_public_constructor_rejects_non_finite_values() -> None:
    with pytest.raises(ValueError, match="lpips.*finite"):
        MetricValues(0.1, math.nan, 1.0, 1.0, 0.5, 0.5)


@pytest.mark.unit
def test_metric_report_row_public_constructor_rejects_unknown_dataset() -> None:
    metrics = MetricValues(0.1, 0.2, 1.0, 2.0, 0.3, 0.4)

    with pytest.raises(ValueError, match="dataset"):
        MetricReportRow(cast(Any, "openvid_validation"), metrics)


@pytest.mark.unit
def test_report_public_constructor_rejects_wrong_row_count_and_order() -> None:
    rows = ThreeDatasetMetricReport.from_mapping(_results()).rows

    with pytest.raises(ValueError, match="exactly four"):
        ThreeDatasetMetricReport(cast(Any, rows[:3]))
    with pytest.raises(ValueError, match="order"):
        ThreeDatasetMetricReport((rows[1], rows[0], rows[2], rows[3]))


@pytest.mark.unit
def test_report_public_constructor_rejects_inconsistent_average() -> None:
    rows = ThreeDatasetMetricReport.from_mapping(_results()).rows
    wrong_average = MetricReportRow("average", MetricValues(0.0, 0.0, 0.0, 0.0, 0.0, 0.0))

    with pytest.raises(ValueError, match="average.*mean"):
        ThreeDatasetMetricReport((*rows[:3], wrong_average))


@pytest.mark.unit
@pytest.mark.parametrize("extension", ["csv", "xlsx"])
def test_report_atomic_replace_failure_preserves_existing_target(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, extension: str
) -> None:
    import tardis.metrics.report as report_module

    destination = tmp_path / f"metrics.{extension}"
    destination.write_bytes(b"existing-report")
    report = ThreeDatasetMetricReport.from_mapping(_results())

    def fail_replace(_source: object, _destination: object) -> None:
        raise OSError("synthetic replace failure")

    monkeypatch.setattr(report_module.os, "replace", fail_replace)

    with pytest.raises(OSError, match="synthetic replace failure"):
        getattr(report, f"write_{extension}")(destination)

    assert destination.read_bytes() == b"existing-report"
    assert not list(tmp_path.glob(f".{destination.name}.*.tmp"))
